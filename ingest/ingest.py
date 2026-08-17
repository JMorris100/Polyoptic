#!/usr/bin/env python3
"""
Polyoptic ingest.

Reads ingest/series.yaml, fetches each series from its source, normalises it
into the series-first schema, and writes data/bundle.json for the front end.

Design notes
------------
* One series = one object. Topics are tags, not folders. A series belongs to
  Housing AND Economy without being duplicated.
* Reference series (population, GDP deflator, GDP) live once, centrally, and
  power the per-person / real-terms / %-of-GDP transforms.
* Nothing is computed at ingest time that could be computed in the browser.
  Store what was published; derive the rest on demand.
* Every series carries geography, year_basis, unit and licence. The front end
  refuses to silently compare across mismatches.
* `publisher` / `publisher_full` are derived here rather than in the front end,
  so the verify page's Source facet and the explorer's citation builder agree
  by construction instead of by two copies of the same lookup table.

Run:
    pip install -r ingest/requirements.txt
    python ingest/ingest.py            # all series
    python ingest/ingest.py --only hpe # one series
    python ingest/ingest.py --dry-run  # fetch and validate, write nothing
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

import requests
import yaml

ROOT = Path(__file__).resolve().parent.parent
CONFIG = ROOT / "ingest" / "series.yaml"
OUT = ROOT / "data" / "bundle.json"
CACHE = ROOT / "ingest" / ".cache"

UA = {"User-Agent": "polyoptic-ingest/0.1 (+https://polyoptic.co)"}
TIMEOUT = 45


# ─────────────────────────────────────────────────────────────
# Publishers
#
# `source` is a release name ("ONS Index of Private Housing Rental
# Prices"), which is the right thing to show on a card but the wrong
# thing to cite: Harvard wants a corporate author, and a facet wants a
# handful of buckets rather than one per release. Both are derived from
# the leading agency in `source`, longest prefix first so "Home Office"
# is not shadowed by a shorter key.
#
# A series can override either with `publisher:` / `publisher_full:` in
# series.yaml when the derived answer is wrong.
# ─────────────────────────────────────────────────────────────

PUBLISHERS: dict[str, str] = {
    "ONS": "Office for National Statistics",
    "Home Office": "Home Office",
    "DfE": "Department for Education",
    "MoJ": "Ministry of Justice",
    "HMRC": "HM Revenue & Customs",
    "FCDO": "Foreign, Commonwealth & Development Office",
    "MHCLG": "Ministry of Housing, Communities & Local Government",
    "HM Treasury": "HM Treasury",
    "HM Land Registry": "HM Land Registry",
    "OBR": "Office for Budget Responsibility",
    "DWP": "Department for Work and Pensions",
    "DHSC": "Department of Health and Social Care",
    "DfT": "Department for Transport",
    "DESNZ": "Department for Energy Security and Net Zero",
    "NHS England": "NHS England",
    "Institute for Government": "Institute for Government",
}


def resolve_publisher(spec: dict) -> tuple[str, str]:
    """Short and full publisher for a series spec.

    Falls back to the first word of `source` so a new agency still gets a
    usable facet before anyone remembers to add it to PUBLISHERS.
    """
    short = spec.get("publisher")
    if not short:
        source = spec.get("source", "")
        matches = [k for k in PUBLISHERS if source.startswith(k)]
        short = max(matches, key=len) if matches else (source.split()[0] if source else "Unknown")
    full = spec.get("publisher_full") or PUBLISHERS.get(short, short)
    return short, full


# ─────────────────────────────────────────────────────────────
# Schema
# ─────────────────────────────────────────────────────────────

@dataclass
class Series:
    id: str
    name: str
    topics: list[str]
    unit: str
    geography: str
    year_basis: str            # calendar | financial | academic
    source: str
    source_url: str
    licence: str
    kind: str                  # money_cash | money_real | count | rate | ratio | index
    start: int
    values: list[float | None]
    publisher: str = ""        # short agency, for facets — "ONS"
    publisher_full: str = ""   # corporate author, for citations
    published: str = ""        # edition year, where one is meaningful; else blank
    areas: bool = False         # true when data/areas/<id>.json exists alongside
    discontinuities: list[dict] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    fetched: str = ""
    source_api: str = ""       # the exact endpoint the values were pulled from

    def validate(self) -> None:
        if self.year_basis not in {"calendar", "financial", "academic"}:
            raise ValueError(f"{self.id}: bad year_basis {self.year_basis!r}")
        if self.kind not in {"money_cash", "money_real", "count", "rate", "ratio", "index"}:
            raise ValueError(f"{self.id}: bad kind {self.kind!r}")
        if not self.values:
            raise ValueError(f"{self.id}: no observations")
        real = [v for v in self.values if v is not None]
        if len(real) < 3:
            raise ValueError(f"{self.id}: only {len(real)} usable observations")


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def get(url: str, params: dict | None = None) -> requests.Response:
    CACHE.mkdir(exist_ok=True)
    r = requests.get(url, params=params, headers=UA, timeout=TIMEOUT)
    r.raise_for_status()
    return r


def get_file(url: str) -> Path:
    """
    Download to ingest/.cache, keyed by URL hash, and reuse it next run.

    Statistical spreadsheets run to tens of megabytes and the publishers put a
    content hash in the path, so a given URL is immutable — caching is safe and
    turns a five-minute rebuild into a five-second one. Delete .cache to refetch.
    """
    import hashlib

    CACHE.mkdir(exist_ok=True)
    # The extension is not always in the path — ONS serves files as
    # /file?uri=/…/table.xlsx — so look anywhere in the URL, last match wins.
    found = re.findall(r"\.(xlsx|xlsm|xls|ods|csv|zip)(?![a-z0-9])", url, re.I)
    suffix = found[-1].lower() if found else "bin"
    path = CACHE / f"{hashlib.sha256(url.encode()).hexdigest()[:16]}.{suffix}"
    if not path.exists() or path.stat().st_size == 0:
        r = requests.get(url, headers=UA, timeout=300)
        r.raise_for_status()
        path.write_bytes(r.content)
    return path


# Ordered, not a set: _period_key needs the index to mean the month number.
MONTHS = ("jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec")


def strip_revision_flags(label: str) -> str:
    """
    Drop the provisional / revised / estimated markers publishers append to
    period labels: 'YE Dec 25 P', 'YE Mar 25 P R', '2024 [R]'.

    Worth doing carefully. Left in, they make a period label unparseable, the
    observation is silently dropped, and a 'last value in the year' series
    quietly reports an older quarter instead — wrong, and wrong in a way that
    looks entirely plausible on a chart.
    """
    return re.sub(r"(?:\s+\[?[PRE]\]?)+$", "", str(label).strip())


def to_year(label: str, basis: str) -> int | None:
    """
    Collapse a period label to a single calendar year for plotting.

    Financial ('2023-24') and academic ('2023/24') years are plotted at their
    STARTING calendar year. This is a real decision with real consequences —
    a 2023/24 academic figure sits at 2023 even though most of it happened in
    2024. The front end surfaces this whenever bases are mixed on one chart.
    """
    label = strip_revision_flags(label)
    if label.endswith(".0") and label[:-2].isdigit():        # 2018.0 from a float cell
        label = label[:-2]
    # CMD time codes look like "Jul-17" — month abbreviation, two-digit year.
    parts = label.split("-")
    if len(parts) == 2 and parts[0][:3].lower() in MONTHS and parts[1].strip().isdigit():
        yy = int(parts[1].strip())
        return 2000 + yy if yy < 50 else 1900 + yy
    for sep in ("-", "/", " to "):
        if sep in label:
            head = label.split(sep)[0].strip()
            if head.isdigit() and len(head) == 4:
                return int(head)
            # Spelt-out ranges: 'Apr 2001 to Mar 2002' is the 2001 financial
            # year, so take the year from the START of the range, never the end.
            found = re.search(r"(?<!\d)(1[89]\d\d|20\d\d)(?!\d)", head)
            if found:
                return int(found.group(1))
    if len(label) >= 4 and label[:4].isdigit():
        return int(label[:4])
    # ONS year-ending labels: "YE Dec 25". Two-digit year, month first, so none
    # of the rules above can see it.
    m = re.match(r"^(?:YE\s+)?([A-Za-z]{3})[a-z]*[\s-](\d{2})$", label, re.I)
    if m and m.group(1).lower() in MONTHS:
        yy = int(m.group(2))
        return 2000 + yy if yy < 50 else 1900 + yy
    # Spreadsheet stock dates: "31 Mar 2026", "as at 30 June 2015". Fall back to
    # the last four-digit number in the label. Deliberately last — anything with
    # a leading year should have matched above.
    tail = re.findall(r"(?<!\d)(1[89]\d\d|20\d\d)(?!\d)", label)
    if tail:
        return int(tail[-1])
    return None


def to_annual(pairs: list[tuple[int, float]], how: str = "mean") -> dict[int, float]:
    """Collapse sub-annual observations to one value per year."""
    buckets: dict[int, list[float]] = {}
    for year, value in pairs:
        buckets.setdefault(year, []).append(value)
    if how == "last":
        return {y: v[-1] for y, v in buckets.items()}
    if how == "sum":
        return {y: sum(v) for y, v in buckets.items()}
    return {y: sum(v) / len(v) for y, v in buckets.items()}


def densify(by_year: dict[int, float]) -> tuple[int, list[float | None]]:
    """Turn a sparse year->value map into (start, [values]) with None for gaps."""
    if not by_year:
        return 0, []
    lo, hi = min(by_year), max(by_year)
    return lo, [by_year.get(y) for y in range(lo, hi + 1)]


# ─────────────────────────────────────────────────────────────
# Adapters
# ─────────────────────────────────────────────────────────────

def fetch_ons_timeseries(cfg: dict) -> dict[int, float]:
    """
    ONS website time-series JSON.

        https://www.ons.gov.uk/<path>/timeseries/<cdid>/<dataset>/data

    This is the pragmatic route for headline economic series — one request,
    everything in it. The trade-off is that the URL encodes the site's own
    taxonomy, so if ONS reorganises a page the URL moves. Pin the CDID in
    config and expect to fix paths occasionally.
    """
    url = f"https://www.ons.gov.uk/{cfg['path'].strip('/')}/timeseries/{cfg['cdid'].lower()}/{cfg['dataset'].lower()}/data"
    cfg["_resolved_url"] = url
    payload = get(url).json()

    granularity = cfg.get("granularity", "years")
    rows = payload.get(granularity) or payload.get("years") or []
    pairs: list[tuple[int, float]] = []
    for row in rows:
        year = to_year(row.get("year") or row.get("date", ""), "calendar")
        raw = row.get("value")
        if year is None or raw in (None, "", ".."):
            continue
        try:
            pairs.append((year, float(str(raw).replace(",", ""))))
        except ValueError:
            continue
    return to_annual(pairs, cfg.get("aggregate", "mean"))


def fetch_ons_beta(cfg: dict) -> dict[int, float]:
    """
    ONS Beta / CMD API — https://api.beta.ons.gov.uk/v1

    Open, no key required. Structure is datasets -> editions -> versions.
    The per-observation endpoint silently returns zero rows for wildcard
    time queries these days, so this goes through each version's bulk CSV
    download instead (the "v4" format: first column is the value, then a
    code/label column pair per dimension).

    `dimensions` in config maps CSV code-column headers to the option code
    to keep, e.g. {administrative-geography: K02000001}. The time value is
    read from the code column named in `time_column`, defaulting to the
    common CMD time headers.
    """
    import csv
    import io

    base = "https://api.beta.ons.gov.uk/v1"
    dataset, edition = cfg["dataset"], cfg.get("edition", "time-series")

    version = cfg.get("version")
    if version in (None, "latest"):
        meta = get(f"{base}/datasets/{dataset}").json()
        latest = meta.get("links", {}).get("latest_version", {})
        version = latest.get("id") or str(latest.get("href", "")).rstrip("/").split("/")[-1]

    vmeta = get(f"{base}/datasets/{dataset}/editions/{edition}/versions/{version}").json()
    csv_href = (vmeta.get("downloads", {}).get("csv") or {}).get("href")
    if not csv_href:
        raise ValueError(f"{dataset} v{version}: no CSV download available")
    cfg["_resolved_url"] = csv_href

    text = get(csv_href).text
    reader = csv.DictReader(io.StringIO(text))
    fields = reader.fieldnames or []

    time_col = cfg.get("time_column")
    if not time_col:
        for cand in ("mmm-yy", "yyyy", "calendar-years", "time", "Time"):
            if cand in fields:
                time_col = cand
                break
    if not time_col:
        raise ValueError(f"{dataset}: can't find a time column in {fields}")

    value_col = fields[0]                       # v4_N — always first
    want = {k: str(v) for k, v in cfg.get("dimensions", {}).items()}

    pairs: list[tuple[int, float]] = []
    for row in reader:
        if any(str(row.get(col, "")).strip() != code for col, code in want.items()):
            continue
        year = to_year(row.get(time_col, ""), "calendar")
        raw = row.get(value_col, "")
        if year is None or raw in (None, "", ".."):
            continue
        try:
            pairs.append((year, float(str(raw).replace(",", ""))))
        except ValueError:
            continue
    return to_annual(pairs, cfg.get("aggregate", "mean"))


def fetch_dfe_ees(cfg: dict) -> dict[int, float]:
    """
    DfE Explore Education Statistics API.

        https://api.education.gov.uk/statistics/v1/data-sets/{id}/query

    POST with a JSON query is the supported production route; GET exists for
    exploration but can't express conditions. Note that not every dataset in
    the EES catalogue is exposed through the API — some are CSV-only, which is
    why `csv` below exists as a fallback source type.
    """
    base = "https://api.education.gov.uk/statistics/v1"
    cfg["_resolved_url"] = f"{base}/data-sets/{cfg['data_set_id']}/query"
    pairs: list[tuple[int, float]] = []
    page = 1
    while True:
        body = {
            "criteria": cfg.get("criteria", {}),
            "indicators": [cfg["indicator"]],
            "page": page,
            "pageSize": 1000,
        }
        r = requests.post(
            f"{base}/data-sets/{cfg['data_set_id']}/query",
            json=body, headers={**UA, "Content-Type": "application/json"}, timeout=TIMEOUT,
        )
        r.raise_for_status()
        payload = r.json()

        for row in payload.get("results", []):
            period = row.get("timePeriod", {})
            year = to_year(str(period.get("period", "")), "academic")
            raw = (row.get("values") or {}).get(cfg["indicator"])
            if year is None or raw in (None, "", "z", "x", "c"):
                continue
            try:
                pairs.append((year, float(str(raw).replace(",", ""))))
            except ValueError:
                continue

        paging = payload.get("paging", {})
        if page >= paging.get("totalPages", 1):
            break
        page += 1
    return to_annual(pairs, cfg.get("aggregate", "mean"))


def fetch_csv(cfg: dict) -> dict[int, float]:
    """
    Generic CSV fallback — MHCLG live tables, OBR databases, anything published
    as a file rather than an API. Give it a URL, a year column and a value
    column. Unglamorous, and you'll use it more than you expect.
    """
    import csv
    import io

    text = get(cfg["url"]).text
    cfg["_resolved_url"] = cfg["url"]
    reader = csv.DictReader(io.StringIO(text))
    pairs: list[tuple[int, float]] = []
    for row in reader:
        for key, want in (cfg.get("filter") or {}).items():
            if str(row.get(key, "")).strip() != str(want):
                break
        else:
            year = to_year(row.get(cfg["year_column"], ""), cfg.get("year_basis", "calendar"))
            raw = row.get(cfg["value_column"], "")
            if year is None or raw in (None, "", ".."):
                continue
            try:
                pairs.append((year, float(str(raw).replace(",", "").replace("£", ""))))
            except ValueError:
                continue
    return to_annual(pairs, cfg.get("aggregate", "mean"))


# ─────────────────────────────────────────────────────────────
# Explore Local Statistics
#
# ONS's subnational service, and the only source here that is an ingest job
# rather than a build: one endpoint, one indicator slug, and a whole history
# comes back as CSV.
#
#   /explore-local-statistics/api/v1/data.csv?indicator=<slug>&geo=<codes>&time=all
#
# `time=all` is the load-bearing part. Leave `time` off and the service
# returns the LATEST PERIOD ONLY, without complaint — a silent one-point
# series. validate() catches that, but only because it insists on three
# observations; don't remove the parameter and trust the shape.
#
# `geo` takes comma-separated area codes and K02000001 is the UK, so the
# national spine and the local detail come out of the same endpoint. The
# same call with no `geo` returns every area at once, which is what the
# area build uses.
#
# Periods are ISO 8601 intervals — '2021-04-01/P1Y' for a financial year,
# '2019-08-01/P1Y' for an academic one. to_year() reads the start date,
# which is already this project's convention for both.
# ─────────────────────────────────────────────────────────────

ELS_API = "https://www.ons.gov.uk/explore-local-statistics/api/v1"
ELS_UK = "K02000001"


def els_data(indicator: str, geo: str | None = None, time: str = "all") -> tuple[str, list[dict]]:
    """Rows from the ELS data endpoint, plus the exact URL they came from."""
    import csv
    import io

    params: dict[str, str] = {"indicator": indicator, "time": time}
    if geo:
        params["geo"] = geo
    r = get(f"{ELS_API}/data.csv", params)
    return r.url, list(csv.DictReader(io.StringIO(r.text)))


def fetch_els(cfg: dict) -> dict[int, float]:
    """
    One ELS indicator for one area, as an annual series.

    Defaults to the UK. `filter:` narrows the multivariate indicators, which
    carry extra `sex` / `age` columns — population-by-age-and-sex is the only
    one in the catalogue today.
    """
    url, rows = els_data(cfg["indicator"], cfg.get("geo", ELS_UK), cfg.get("time", "all"))
    cfg["_resolved_url"] = url

    pairs: list[tuple[int, float]] = []
    for row in rows:
        for key, want in (cfg.get("filter") or {}).items():
            if str(row.get(key, "")).strip() != str(want):
                break
        else:
            year = to_year(row.get("period", ""), cfg.get("year_basis", "calendar"))
            raw = row.get("value", "")
            if year is None or raw in (None, "", ":", ".."):
                continue
            try:
                pairs.append((year, float(str(raw).replace(",", ""))))
            except ValueError:
                continue
    if not pairs:
        raise ValueError(f"ELS returned no usable rows for {cfg['indicator']!r} at {cfg.get('geo', ELS_UK)}")
    return to_annual(pairs, cfg.get("aggregate", "mean"))


# ─────────────────────────────────────────────────────────────
# Spreadsheet adapter
#
# Most of British government statistics is not an API. The Home Office, MoJ,
# HMRC, FCDO and the ONS crime team all publish .xlsx / .ods workbooks and
# nothing else, so this adapter is the difference between a catalogue that can
# cover borders, crime and tax and one that can only cover the economy.
#
# Two layouts, because publishers use both:
#   long — one row per observation, dimensions in columns (Home Office style)
#   wide — one row per measure, years across the columns (ONS table style)
# ─────────────────────────────────────────────────────────────

def _clean(v) -> str:
    """Header and label text as published, minus the whitespace and footnote junk."""
    s = "" if v is None else str(v)
    # The same Home Office column is 'Date (as at…)' on one sheet and
    # 'Date (as at...)' on the next, so normalise the ellipsis rather than
    # asking the catalogue to guess which spelling a given workbook used.
    s = s.replace("\xa0", " ").replace("’", "'").replace("…", "...")
    s = re.sub(r"\s*\[(note|footnote)[^\]]*\]", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def _to_number(raw) -> float | None:
    """Spreadsheet cell to float, or None for the many flavours of 'no data'."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    s = str(raw).strip().replace(",", "").replace("£", "").replace("%", "")
    if s in {"", "..", ":", "-", "–", "z", "x", "c", "N/A", "n/a", "[z]", "[x]", "[c]", "[low]"}:
        return None
    if s.startswith("(") and s.endswith(")"):                 # (1,234) is negative
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _read_sheet(path: Path, sheet: str | None) -> list[list]:
    """
    Read one worksheet into a list of rows. Handles .xlsx and .ods.

    Parsed rows are cached next to the download. The Home Office crime-outcomes
    workbooks are 700,000 rows and take about 90 seconds each to parse, and
    several series read the same sheet — so without this, adding one indicator
    adds minutes to every build.
    """
    import pickle

    cached = path.with_name(f"{path.stem}.{(sheet or 'first').replace(' ', '_')}.pkl")
    if cached.exists() and cached.stat().st_mtime >= path.stat().st_mtime:
        try:
            return pickle.loads(cached.read_bytes())
        except Exception:                     # noqa: BLE001 — a bad cache is not fatal
            cached.unlink(missing_ok=True)

    rows = _parse_sheet(path, sheet)
    try:
        cached.write_bytes(pickle.dumps(rows, protocol=pickle.HIGHEST_PROTOCOL))
    except OSError:
        pass
    return rows


def _pick_sheet(names: list[str], sheet: str) -> str:
    """
    Resolve a sheet name, exactly if possible and loosely if not.

    Publishers put the period in the tab name and are not consistent about it —
    the same table is 'Outcomes_open_data_2022_23' one year and 'Outcomes open
    data 2025_26' the next. Matching on the normalised name lets one catalogue
    entry span every year's workbook.
    """
    if sheet in names:
        return sheet
    flat = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())
    loose = [n for n in names if flat(sheet) in flat(n)]
    if loose:
        return loose[0]
    raise ValueError(f"no sheet {sheet!r}; workbook has {names}")


def _parse_sheet(path: Path, sheet: str | None) -> list[list]:
    if path.suffix == ".ods":
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P

        doc = load(str(path))
        tables = doc.spreadsheet.getElementsByType(Table)
        names = [t.getAttribute("name") for t in tables]
        table = tables[0] if sheet is None else tables[names.index(_pick_sheet(names, sheet))]

        rows: list[list] = []
        for tr in table.getElementsByType(TableRow):
            row: list = []
            for tc in tr.getElementsByType(TableCell):
                repeat = int(tc.getAttribute("numbercolumnsrepeated") or 1)
                value = tc.getAttribute("value")
                if value is None:
                    value = "\n".join(str(p) for p in tc.getElementsByType(P)) or None
                else:
                    value = float(value)
                # ODS pads rows out to the sheet width with a huge repeat count.
                row.extend([value] * min(repeat, 512))
            while row and row[-1] is None:
                row.pop()
            rows.append(row)
        return rows

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_pick_sheet(wb.sheetnames, sheet)] if sheet else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _find_header(rows: list[list], must_have: list[str]) -> int:
    """Locate the header row by the column names it has to contain."""
    wanted = {w.lower() for w in must_have if w}
    for i, row in enumerate(rows[:80]):
        have = {_clean(c).lower() for c in row if c is not None}
        if wanted <= have:
            return i
    raise ValueError(f"no header row containing {sorted(wanted)} in the first 80 rows")


def _period_key(label: str) -> tuple:
    """
    Sort key putting sub-annual periods in calendar order.

    This exists because 'last' means the last *quarter*, and the published
    labels ('31 Dec 2015', '30 Jun 2015') sort into the wrong order as plain
    strings — which would quietly hand you March instead of December.
    """
    s = re.sub(r"^YE\s+", "", strip_revision_flags(label), flags=re.I)   # "YE Dec 25 P" -> "Dec 25"
    m = re.match(r"^(\d{4})\s*Q([1-4])$", s)                      # 2018 Q1
    if m:
        return (int(m.group(1)), int(m.group(2)) * 3)
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})[a-z]*\s+(\d{4})$", s)   # 31 Dec 2015
    if m and m.group(2).lower() in MONTHS:
        return (int(m.group(3)), MONTHS.index(m.group(2).lower()) + 1, int(m.group(1)))
    m = re.match(r"^([A-Za-z]{3})[a-z]*[- ](\d{2,4})$", s)        # Jul-17, Jul 2017
    if m and m.group(1).lower() in MONTHS:
        yy = int(m.group(2))
        yy = yy if yy > 100 else (2000 + yy if yy < 50 else 1900 + yy)
        return (yy, MONTHS.index(m.group(1).lower()) + 1)
    return (to_year(s, "calendar") or 0, s)


def _matches(cell, want) -> bool:
    """Filter test. A list means membership; a scalar means equality."""
    got = _clean(cell)
    if isinstance(want, (list, tuple, set)):
        return got in {_clean(w) for w in want}
    return got == _clean(want)


def fetch_spreadsheet(cfg: dict) -> dict[int, float]:
    """
    Published .xlsx / .ods workbook -> annual series.

    long layout (default)
        year_column, value_column, optional period_column and filter.
        Values are summed within each period, then collapsed to a year with
        `aggregate`. The two stages matter: a quarterly *stock* like the asylum
        backlog must be summed across its breakdown rows but taken as the LAST
        quarter of the year, never summed across quarters.

        `rate` turns two filters into a percentage — numerator rows over all
        filtered rows — for genuine rate measures (grant rates, charge rates)
        where the publisher only gives you the counts.

    wide layout
        Years run across a header row; one data row holds the measure. Give
        `row_match` (the row label) and `label_column`.
    """
    # Some publishers put one file per year (crime outcomes), so a series can
    # span several workbooks. Same shape in each; merge the years.
    urls = cfg.get("urls") or [cfg["url"]]
    cfg["_resolved_url"] = urls[0] if len(urls) == 1 else f"{urls[0]} (+{len(urls) - 1} more)"

    merged: dict[int, float] = {}
    for url in urls:
        rows = _read_sheet(get_file(url), cfg.get("sheet"))
        merged.update(_wide(cfg, rows) if cfg.get("layout", "long") == "wide"
                      else _long(cfg, rows))

    # `scale` exists because publishers disagree about whether a percentage is
    # 15.5 or 0.155, and the unit string in the catalogue has to be true.
    offset, scale = cfg.get("year_offset", 0), cfg.get("scale", 1)
    if offset or scale != 1:
        merged = {y + offset: v * scale for y, v in merged.items()}
    return merged


def _long(cfg: dict, rows: list[list]) -> dict[int, float]:
    """One row per observation, dimensions in columns."""
    year_col, value_col = cfg["year_column"], cfg["value_column"]
    header_row = cfg.get("header_row")
    header_row = header_row - 1 if header_row else _find_header(rows, [year_col, value_col])
    header = [_clean(c) for c in rows[header_row]]

    def index_of(name: str) -> int:
        try:
            return header.index(_clean(name))
        except ValueError:
            raise ValueError(f"no column {name!r}; have {[h for h in header if h]}") from None

    yi, vi = index_of(year_col), index_of(value_col)
    pi = index_of(cfg["period_column"]) if cfg.get("period_column") else None
    keep = {index_of(k): v for k, v in (cfg.get("filter") or {}).items()}
    numer = {index_of(k): v for k, v in (cfg.get("rate") or {}).items()}

    # (year, period) -> [denominator total, numerator total]
    buckets: dict[tuple[int, str], list[float]] = {}
    for row in rows[header_row + 1:]:
        if len(row) <= max(yi, vi, pi or 0, *keep, *numer, 0):
            row = list(row) + [None] * 64
        if any(not _matches(row[i], want) for i, want in keep.items()):
            continue
        year = to_year(_clean(row[yi]), "calendar")
        value = _to_number(row[vi])
        if year is None or value is None:
            continue
        slot = buckets.setdefault((year, _clean(row[pi]) if pi is not None else ""), [0.0, 0.0])
        slot[0] += value
        if numer and all(_matches(row[i], want) for i, want in numer.items()):
            slot[1] += value

    if not buckets:
        raise ValueError("no rows matched the filter — check the option spellings")

    ordered = sorted(buckets.items(), key=lambda kv: (kv[0][0], _period_key(kv[0][1])))
    how = cfg.get("aggregate", "sum")
    if numer:
        # Collapse numerator and denominator to the year SEPARATELY, then
        # divide. Averaging four quarterly rates instead would weight a quiet
        # quarter the same as a busy one and quietly misstate the year.
        den = to_annual([(y, d) for (y, _), (d, _n) in ordered], how)
        num = to_annual([(y, n) for (y, _), (_d, n) in ordered], how)
        return {y: 100.0 * num[y] / den[y] for y in den if den.get(y)}
    return to_annual([(y, d) for (y, _), (d, _n) in ordered], how)


def _wide(cfg: dict, rows: list[list]) -> dict[int, float]:
    """
    Years across the top, measures down the side.

    `row_filter` narrows by other columns on the same row — MoJ's prison tables
    repeat every nationality row once per sex, so matching the label alone would
    silently pick whichever block happens to come first.

    `column_match` keeps only year columns whose header contains that text. ONS
    outcome tables run a *rolling* quarterly series across the columns ('Year
    ending Mar 2015', 'Year ending Jun 2015', …); without this the last quarter
    in each calendar year wins and the financial-year figure is lost.
    """
    label_col = cfg.get("label_column", 0)
    want = _clean(cfg["row_match"])
    row_filter = {int(k): v for k, v in (cfg.get("row_filter") or {}).items()}

    year_row = cfg.get("year_header_row")
    if year_row:
        year_row -= 1
    else:                                    # the row with the most parseable years wins
        year_row = max(
            range(min(len(rows), 40)),
            key=lambda i: sum(to_year(_clean(c), "calendar") is not None
                              for c in rows[i] if _clean(c)),
        )

    # ONS wide tables end with derived columns — '…compared with previous year
    # % change', '…Significance' — whose headers still contain a year. Left in,
    # a -1.36% change lands in the series as if it were 5.2 million offences.
    derived = re.compile(r"%\s*change|percentage change|significance|unweighted base", re.I)

    column_match = _clean(cfg.get("column_match", "")).lower()
    years: dict[int, int | None] = {}
    for i, cell in enumerate(rows[year_row]):
        header = _clean(cell)
        if derived.search(header):
            continue
        if column_match and column_match not in header.lower():
            continue
        years[i] = to_year(header, "calendar")

    for row in rows[year_row + 1:]:
        if label_col >= len(row) or _clean(row[label_col]) != want:
            continue
        if any(i >= len(row) or not _matches(row[i], v) for i, v in row_filter.items()):
            continue
        out: dict[int, float] = {}
        for i, cell in enumerate(row):
            year, value = years.get(i), _to_number(cell)
            # First column for a year wins. These tables run left-to-right in
            # time and put derived columns last, so anything arriving second is
            # a restatement or a rounding variant, not a newer observation.
            if year is not None and value is not None and i != label_col and year not in out:
                out[year] = value
        if out:
            return out
    labels = [_clean(r[label_col]) for r in rows[year_row + 1:] if label_col < len(r)]
    raise ValueError(f"no row labelled {want!r}; have {[l for l in labels if l][:25]}")


ADAPTERS = {
    "ons_timeseries": fetch_ons_timeseries,
    "ons_beta": fetch_ons_beta,
    "dfe_ees": fetch_dfe_ees,
    "csv": fetch_csv,
    "els": fetch_els,
    "spreadsheet": fetch_spreadsheet,
}


# ─────────────────────────────────────────────────────────────
# ELS catalogue generator
#
# The ELS metadata endpoint describes all 109 indicators in enough detail to
# write most of a series.yaml entry — label, unit, source, coverage, and the
# publisher's own caveats. Generating those blocks rather than hand-keying
# them is the difference between "a geography roadmap" and an afternoon.
#
# It writes nothing. It prints candidate blocks for a human to read, edit and
# paste, because the parts it cannot get right are the parts that matter: the
# permalink id, the topic tags, and whether a caveat is worth a note.
# ─────────────────────────────────────────────────────────────

# ELS names its own topics; this catalogue's tags are narrower and older.
# Sub-topic wins where it is more specific than the topic above it.
ELS_TOPICS = {
    "economy": "economy", "health and wellbeing": "health",
    "education and skills": "education", "population": "population",
    "housing": "housing", "environment": "environment", "crime": "crime",
    "connectivity": "connectivity",
}
ELS_SUBTOPICS = {
    "travel and transport": "transport", "digital connectivity": "digital",
    "access to amenities": "amenities", "culture and heritage": "culture",
    "pay and income": "economy", "employment": "economy",
    "business": "economy", "productivity": "economy", "trade": "economy",
}
ELS_COUNTRIES = {
    "ENSW": "UK", "ESW": "Great Britain", "EW": "England and Wales",
    "E": "England", "S": "Scotland", "W": "Wales", "N": "Northern Ireland",
    "ENW": "England, Wales and Northern Ireland",
}
ELS_BASIS = {"financial-year": "financial", "academic-year": "academic"}


def els_kind(ind: dict) -> str:
    unit = (ind.get("unit") or "").strip()
    if ind.get("suffix") == "%" or unit == "%" or "%" in unit:
        return "rate"
    if ind.get("prefix") == "£" or unit.startswith("£"):
        return "money_cash"
    if unit.lower().startswith(("rate per", "count per")):
        return "rate"
    if unit.lower() in {"years", "yrs", "mins", "minutes"} or unit.lower().startswith("score"):
        return "ratio"
    return "count"


def els_unit(ind: dict) -> str:
    """
    A unit string that reads on a chart axis, not a database column.

    `unit` is the column heading ('Persons'), `subText` the human gloss
    ('persons per square kilometre'). The gloss wins where there is one —
    it is the difference between an axis that says what the number means
    and one that says what type it is.
    """
    unit = (ind.get("unit") or "").strip()
    sub = re.sub(r"^in\s+", "", (ind.get("subText") or "").strip())
    if sub:
        return f"% {sub}" if unit == "%" and "%" not in sub else sub
    return unit or (ind.get("measure") or "").strip() or "value"


def els_notes(ind: dict) -> list[str]:
    """The publisher's caveats, plus the ones implied by the metadata."""
    notes: list[str] = []
    fmt = ind.get("periodFormat") or ""
    m = re.match(r"^(\d+)-year$", fmt)
    if m and m.group(1) != "1":
        notes.append(
            f"Pooled {m.group(1)}-year estimates, plotted at the first year of each band — "
            f"a point here summarises {m.group(1)} years, not one")
    titled = NATION_SUFFIX.search(ind.get("label", ""))
    if titled and "K02" in (ind.get("geography") or {}).get("types", []):
        notes.append(
            f"ONS titles this indicator {titled.group(1)} but serves the top-level row as "
            f"'United Kingdom'. Published here on the narrower claim in the title")
    if ind.get("experimentalStatistic"):
        notes.append("Published by ONS as an experimental statistic — the method is still under development")
    srcs = ind.get("source") or []
    if len(srcs) > 1:
        who = ", ".join(s["name"] for s in srcs)
        notes.append(f"Assembled by ONS from more than one collection ({who}), so the nations are not always on identical definitions")
    for c in ind.get("caveats") or []:
        for para in str(c).split("\n\n"):
            para = " ".join(para.split())
            if para:
                notes.append(para)
    return notes


NATION_SUFFIX = re.compile(
    r"\s*\((Great Britain|England and Wales|England|Wales|Scotland|Northern Ireland)\)\s*$")


def els_geography(ind: dict) -> str:
    """
    Coverage, taking the indicator's own title over its country list.

    These disagree, and not harmlessly: 'Employment rate (Great Britain)'
    lists all four countries and serves a row labelled 'United Kingdom' at
    K02000001. The title is the narrower claim, so it is the one to publish
    — els_notes() records the discrepancy rather than hiding it.
    """
    titled = NATION_SUFFIX.search(ind.get("label", ""))
    if titled:
        return titled.group(1)
    countries = "".join(sorted((ind.get("geography") or {}).get("countries", [])))
    return ELS_COUNTRIES.get(countries, countries or "UK")


def _yaml_str(s: str) -> str:
    return '"' + str(s).replace("\\", "\\\\").replace('"', '\\"') + '"'


def els_block(ind: dict, geo: str) -> str:
    """One series.yaml entry, in the house style, ready to review."""
    label = ind["label"]
    # Labels carry their own coverage note — 'Employment rate (Great Britain)'.
    # geography: already says that, so take it out of the name.
    name = re.sub(r"\s*\((?:Great Britain|England|Wales|Scotland|Northern Ireland)\)\s*$", "", label)
    topic = ELS_SUBTOPICS.get(ind.get("subTopic") or "") or ELS_TOPICS.get(ind.get("topic") or "", "economy")

    srcs = ind.get("source") or [{}]
    lead = srcs[0].get("name", "")
    short = next((k for k, v in PUBLISHERS.items() if v == lead), None)

    lines = [
        f"  - id: {ind['slug']}",
        f"    name: {name}",
        f"    topics: [{topic}]",
        f"    unit: {_yaml_str(els_unit(ind))}",
        f"    kind: {els_kind(ind)}",
        f"    geography: {els_geography(ind)}",
        f"    year_basis: {ELS_BASIS.get(ind.get('periodFormat') or '', 'calendar')}",
        "    source: ONS Explore Local Statistics",
        f"    source_url: {srcs[0].get('href', '')}",
    ]
    if short:
        lines.append(f"    publisher: {short}")
    elif lead:
        lines.append(f"    publisher: {lead.split()[0]}")
        lines.append(f"    publisher_full: {_yaml_str(lead)}")
    lines += [
        "    source_type: els",
        "    fetch:",
        f"      indicator: {ind['slug']}",
    ]
    if geo != ELS_UK:
        lines.append(f"      geo: {geo}")
    if (ind.get("frequency") or "annual") != "annual":
        lines.append("      aggregate: mean          # sub-annual source — check this is the right collapse")
    if ind.get("isMultivariate"):
        lines.append("      filter: {}               # multivariate: needs sex / age pinning")
    notes = els_notes(ind)
    if notes:
        lines.append("    notes:")
        lines += [f"      - {_yaml_str(n)}" for n in notes]
    return "\n".join(lines)


def els_catalogue(geo: str, min_span: int) -> int:
    indicators = get(f"{ELS_API}/metadata/indicators").json()
    wanted = geo[:3]
    out, skipped = [], []
    for ind in sorted(indicators, key=lambda i: (i.get("topic") or "", i["slug"])):
        levels = (ind.get("geography") or {}).get("types", [])
        domain = ind.get("periodDomain") or []
        span = 0
        if len(domain) >= 2:
            years = [to_year(d, "calendar") for d in (domain[0], domain[-1])]
            span = (years[1] - years[0] + 1) if all(y is not None for y in years) else 0
        if wanted not in levels:
            skipped.append((ind["slug"], f"no {wanted} area"))
        elif not ind.get("hasTimeseries"):
            skipped.append((ind["slug"], "no time series"))
        elif span < min_span:
            skipped.append((ind["slug"], f"{span}-year span"))
        else:
            out.append(els_block(ind, geo))

    print(f"# {len(out)} ELS indicators at {geo}, {min_span}+ year span.")
    print("# Review before pasting: `id` is a permalink token so keep it short,")
    print("# `topics` are this catalogue's tags, and not every caveat needs a note.\n")
    print("\n\n".join(out))
    print(f"\n# skipped {len(skipped)}:", file=sys.stderr)
    for slug, why in skipped:
        print(f"#   {slug:<52} {why}", file=sys.stderr)
    return 0


# ─────────────────────────────────────────────────────────────
# Area build
#
# The same ELS call without `geo` returns every area at once — around 400
# of them, four nations, five levels — so a series that is already in the
# catalogue gains its local detail for one extra request.
#
# It does not go in bundle.json. The bundle is what every visitor downloads
# before they see a chart, and the area data is thirty times its size for a
# view most visits never open. One file per series under data/areas/, plus a
# small index the front end loads when someone actually asks for areas.
#
# Levels come from the code prefix rather than from the row, because ONS's
# own area-level naming varies between the API and the published files while
# the coding scheme itself is stable and documented.
# ─────────────────────────────────────────────────────────────

AREAS_DIR = ROOT / "data" / "areas"

ELS_LEVELS = {
    "K02": "uk", "K03": "uk", "K04": "uk", "K99": "uk",
    "E92": "country", "W92": "country", "S92": "country", "N92": "country",
    "E12": "region", "E13": "region",
    # Combined authorities in England; the city-region and growth-deal bodies
    # in Wales and Northern Ireland are the nearest equivalent, so they share
    # the level rather than getting one row each of their own.
    "E47": "cauth", "E48": "cauth", "W42": "cauth", "N34": "cauth",
    "E10": "county", "E11": "county",
    "E06": "la", "E08": "la", "E09": "la", "W06": "la", "S12": "la", "N09": "la",
    "E07": "district",
}

# ITL — the UK's own replacement for NUTS, used by the regional accounts.
# The level is the code's LENGTH, not its prefix: TLC, TLC3, TLC31.
ITL_LEVELS = {3: "itl1", 4: "itl2", 5: "itl3"}

LEVEL_NAMES = {
    "uk": "UK and nation groupings", "country": "Countries", "region": "Regions",
    "cauth": "Combined and city-region authorities", "county": "Counties",
    "la": "Local authorities", "district": "Districts",
    "itl1": "ITL1 regions", "itl2": "ITL2 areas", "itl3": "ITL3 areas",
}

# Coarsest first — the front end offers levels in this order.
LEVEL_ORDER = ["uk", "country", "region", "itl1", "cauth", "county",
               "itl2", "la", "district", "itl3", "other"]


def els_level(code: str) -> str:
    if code.startswith("TL"):
        return ITL_LEVELS.get(len(code), "other")
    return ELS_LEVELS.get(code[:3], "other")


def build_areas(specs: list[dict], only: str | None = None) -> dict:
    """Write data/areas/<id>.json for every ELS series, and an index of them."""
    els = [s for s in specs if s.get("source_type") == "els"]
    if only:
        els = [s for s in els if s["id"] == only]
    if not els:
        return {}

    AREAS_DIR.mkdir(parents=True, exist_ok=True)

    registry: dict[str, dict] = {}
    index: dict[str, dict] = {}

    for spec in els:
        sid, slug = spec["id"], spec["fetch"]["indicator"]
        try:
            url, rows = els_data(slug, None, "all")
        except Exception as exc:                          # noqa: BLE001
            print(f"  FAIL areas:{sid} — {exc}", file=sys.stderr)
            continue

        by_area: dict[str, dict[int, float]] = {}
        for row in rows:
            for key, want in (spec["fetch"].get("filter") or {}).items():
                if str(row.get(key, "")).strip() != str(want):
                    break
            else:
                code = (row.get("areacd") or "").strip()
                year = to_year(row.get("period", ""), spec.get("year_basis", "calendar"))
                raw = row.get("value", "")
                # ELS occasionally serves a row coded literally 'na'. An area
                # with no code cannot be identified, mapped or linked to, so
                # it is dropped rather than shown under its name alone.
                if not code or code == "na" or year is None or raw in (None, "", ":", ".."):
                    continue
                try:
                    value = float(str(raw).replace(",", ""))
                except ValueError:
                    continue
                # Rounded only enough to keep float noise ('533.9000000000001')
                # out of the file. NOT to the indicator's decimalPlaces: that
                # is a display hint, and these files get downloaded and joined,
                # so they carry the figure as published and let the chart
                # decide how many digits to show.
                by_area.setdefault(code, {})[year] = round(value, 6)
                registry.setdefault(code, {
                    "n": (row.get("areanm") or code).strip(),
                    "l": els_level(code),
                })

        if not by_area:
            print(f"  FAIL areas:{sid} — no rows", file=sys.stderr)
            continue

        lo = min(y for v in by_area.values() for y in v)
        hi = max(y for v in by_area.values() for y in v)
        span = range(lo, hi + 1)
        values = {code: [v.get(y) for y in span] for code, v in sorted(by_area.items())}

        payload = {"id": sid, "indicator": slug, "start": lo,
                   "source_api": url, "values": values}
        (AREAS_DIR / f"{sid}.json").write_text(
            json.dumps(payload, separators=(",", ":")), encoding="utf-8")

        seen = {registry[c]["l"] for c in by_area}
        levels = [lv for lv in LEVEL_ORDER if lv in seen]
        index[sid] = {"start": lo, "end": hi, "areas": len(by_area), "levels": levels}
        print(f"  ok   areas:{sid:<12} {lo}–{hi}  {len(by_area):>4} areas")

    (AREAS_DIR / "index.json").write_text(json.dumps({
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "levelNames": LEVEL_NAMES,
        "levelOrder": LEVEL_ORDER,
        "areas": dict(sorted(registry.items())),
        "series": index,
    }, separators=(",", ":")), encoding="utf-8")
    return index


# ─────────────────────────────────────────────────────────────
# Build
# ─────────────────────────────────────────────────────────────

def build_series(spec: dict) -> Series:
    adapter = ADAPTERS[spec["source_type"]]
    by_year = adapter(spec["fetch"])
    start, values = densify(by_year)
    publisher, publisher_full = resolve_publisher(spec)
    s = Series(
        id=spec["id"], name=spec["name"], topics=spec["topics"], unit=spec["unit"],
        geography=spec["geography"], year_basis=spec["year_basis"],
        source=spec["source"], source_url=spec.get("source_url", ""),
        licence=spec.get("licence", "OGL v3.0"), kind=spec["kind"],
        start=start, values=values,
        publisher=publisher, publisher_full=publisher_full,
        published=str(spec.get("published", "")),
        discontinuities=spec.get("discontinuities", []),
        notes=spec.get("notes", []),
        fetched=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        source_api=spec["fetch"].get("_resolved_url", ""),
    )
    s.validate()
    return s


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", help="build a single series by id")
    ap.add_argument("--dry-run", action="store_true", help="fetch and validate, write nothing")
    ap.add_argument("--els-catalogue", action="store_true",
                    help="print candidate series.yaml blocks for the ONS Explore Local Statistics indicators")
    ap.add_argument("--els-geo", default=ELS_UK, help="area code for --els-catalogue (default: UK)")
    ap.add_argument("--els-min-span", type=int, default=6, help="shortest run of years worth offering (default: 6)")
    ap.add_argument("--no-areas", action="store_true",
                    help="skip the per-area files under data/areas/")
    ap.add_argument("--areas-only", action="store_true",
                    help="rebuild data/areas/ and leave bundle.json alone")
    args = ap.parse_args()

    if args.els_catalogue:
        return els_catalogue(args.els_geo, args.els_min_span)

    config = yaml.safe_load(CONFIG.read_text(encoding="utf-8"))
    specs = config["series"]
    if args.only:
        specs = [s for s in specs if s["id"] == args.only]
        if not specs:
            print(f"no series with id {args.only!r}", file=sys.stderr)
            return 1

    if args.areas_only:
        return 0 if build_areas(specs, args.only) else 1

    built, failed = [], []
    for spec in specs:
        try:
            s = build_series(spec)
            built.append(s)
            span = f"{s.start}–{s.start + len(s.values) - 1}"
            print(f"  ok   {s.id:<16} {span}  {len([v for v in s.values if v is not None]):>4} obs")
        except Exception as exc:                      # noqa: BLE001 — report and continue
            failed.append((spec["id"], exc))
            print(f"  FAIL {spec['id']:<16} {type(exc).__name__}: {exc}", file=sys.stderr)

    refs = {}
    for name, spec in config.get("references", {}).items():
        try:
            start, values = densify(ADAPTERS[spec["source_type"]](spec["fetch"]))
            refs[name] = {"start": start, "values": values,
                          "source_api": spec["fetch"].get("_resolved_url", "")}
            print(f"  ok   ref:{name}")
        except Exception as exc:                      # noqa: BLE001
            print(f"  FAIL ref:{name} — {exc}", file=sys.stderr)
            print("       transforms depending on this reference will be unavailable", file=sys.stderr)

    ok = {s.id for s in built}

    # Areas are built after the series so a series that failed its national
    # fetch cannot advertise local detail the explorer would then load.
    areas: dict[str, dict] = {}
    if not (args.no_areas or args.dry_run):
        areas = build_areas([s for s in specs if s["id"] in ok], args.only)
        for s in built:
            s.areas = s.id in areas

    bundle = {
        "meta": {
            "provenance": "live" if not failed else "partial",
            "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "failed": [fid for fid, _ in failed],
        },
        "eras": config["eras"],
        "pms": config["pms"],
        # Modes are emitted with only the series that actually built, so a
        # failed fetch can never leave the front end asking for a missing id.
        "modes": {
            key: {**mode, "series": [sid for sid in mode.get("series", []) if sid in ok]}
            for key, mode in config.get("modes", {}).items()
        },
        "refs": refs,
        "series": [asdict(s) for s in built],
    }

    if args.dry_run:
        print(f"\ndry run: {len(built)} built, {len(failed)} failed, nothing written")
        return 1 if failed else 0

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(bundle, separators=(",", ":")), encoding="utf-8")
    size = OUT.stat().st_size / 1024
    print(f"\nwrote {OUT.relative_to(ROOT)} — {len(built)} series, {size:.0f} KB")
    if failed:
        print(f"{len(failed)} failed; bundle marked 'partial'", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
